"""
EXCEL DASHBOARDS GENERATOR
Creates 3 interactive library dashboards in Excel format
Author: Library Analytics Team
Date: 2024
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, timedelta

# ==================== CONFIGURATION ====================
OUTPUT_FOLDER = "library_dashboards"
DASHBOARD_TITLES = {
    'executive': "Library Executive Dashboard",
    'department': "Department Usage Dashboard", 
    'operational': "Operational Metrics Dashboard"
}

# Color schemes for each dashboard
COLOR_PALETTES = {
    'executive': {
        'primary': '2E5A88',      # Navy blue
        'secondary': '4A90E2',    # Light blue
        'accent': '7ED321',       # Green
        'text': 'FFFFFF',         # White
        'alert': 'D0021B'         # Red
    },
    'department': {
        'primary': '8B72BE',      # Purple
        'secondary': '50E3C2',    # Teal
        'accent': 'F5A623',       # Orange
        'text': 'FFFFFF',
        'alert': '9013FE'         # Violet
    },
    'operational': {
        'primary': 'D0021B',      # Red
        'secondary': 'F5A623',    # Orange
        'accent': '417505',       # Dark green
        'text': 'FFFFFF',
        'alert': 'FF6F61'         # Coral
    }
}

# ==================== DATA GENERATION ====================
def generate_library_data():
    """Generate comprehensive library data for all dashboards"""
    
    # Current date for time series
    today = datetime.now()
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    # 1. EXECUTIVE DASHBOARD DATA
    executive_data = {
        'KPI_Summary': pd.DataFrame({
            'Metric': [
                'Total Library Visits', 
                'Digital Resource Downloads', 
                'Physical Book Checkouts',
                'Study Room Bookings (hrs)',
                'Average Visit Duration (min)',
                'User Satisfaction Score (%)',
                'Collection Utilization Rate (%)',
                'Return On Time Rate (%)'
            ],
            'Current_Value': [12456, 18923, 4567, 1876.5, 48, 92.5, 76.4, 88.3],
            'Previous_Value': [11890, 17234, 4321, 1654.2, 45, 89.2, 71.8, 85.7],
            'Target': [13000, 20000, 5000, 2000, 45, 90, 75, 85],
            'Trend': ['↑', '↑', '↑', '↑', '↑', '↑', '↑', '↑'],
            'Status': ['On Track', 'On Track', 'Below Target', 'Below Target', 
                      'Above Target', 'Above Target', 'Above Target', 'Above Target']
        }),
        
        'Monthly_Trends': pd.DataFrame({
            'Month': months,
            'Visitors': [2456, 2678, 2890, 2789, 2912, 2845, 2345, 2567, 3012, 3123, 2876, 2765],
            'Digital_Downloads': [1234, 1456, 1678, 1543, 1621, 1589, 1345, 1421, 1789, 1892, 1674, 1567],
            'Book_Checkouts': [342, 378, 415, 389, 402, 398, 287, 312, 456, 478, 423, 401],
            'Room_Bookings': [156, 167, 189, 178, 192, 187, 145, 156, 201, 215, 198, 184]
        }),
        
        'Resource_Categories': pd.DataFrame({
            'Category': ['Textbooks', 'Research Journals', 'Reference Books', 
                        'Fiction/Literature', 'Multimedia', 'Theses', 'E-Books', 'Magazines'],
            'Usage_Count': [1245, 987, 876, 654, 432, 321, 2345, 198],
            'Percentage': [28.5, 22.6, 20.1, 15.0, 9.9, 7.4, 53.8, 4.5]
        }),
        
        'Peak_Hours': pd.DataFrame({
            'Time_Slot': ['8-9 AM', '9-10 AM', '10-11 AM', '11-12 PM', '12-1 PM', 
                         '1-2 PM', '2-3 PM', '3-4 PM', '4-5 PM', '5-6 PM'],
            'Average_Visitors': [45, 89, 156, 234, 198, 167, 189, 156, 123, 78],
            'Category': ['Low', 'Medium', 'High', 'Peak', 'High', 'Medium', 'High', 'Medium', 'Low', 'Low']
        })
    }
    
    # 2. DEPARTMENT DASHBOARD DATA
    departments = ['Computer Science', 'Engineering', 'Business', 'Arts & Humanities', 
                  'Sciences', 'Social Sciences', 'Medicine', 'Law']
    
    department_data = {
        'Department_Comparison': pd.DataFrame({
            'Department': departments,
            'Total_Users': [456, 345, 234, 123, 234, 345, 123, 234],
            'Book_Checkouts': [1245, 987, 678, 345, 456, 567, 234, 456],
            'Digital_Downloads': [5678, 4321, 2345, 1234, 2345, 3456, 1234, 2345],
            'Research_Requests': [45, 34, 23, 12, 23, 34, 12, 23],
            'Satisfaction_Score': [94, 88, 92, 86, 90, 91, 87, 89]
        }),
        
        'Department_Trends': pd.DataFrame({
            'Month': months[:6] * 3,
            'Department': ['Computer Science']*6 + ['Engineering']*6 + ['Business']*6,
            'Activity_Index': [145, 156, 167, 145, 156, 178, 98, 105, 112, 98, 105, 117, 
                              67, 72, 78, 67, 72, 81, 45, 48, 52, 45, 48, 54],
            'Resource_Usage': [234, 256, 278, 234, 256, 289, 156, 167, 178, 156, 167, 189,
                              98, 105, 112, 98, 105, 118, 67, 72, 78, 67, 72, 80]
        }),
        
        'Top_Resources_by_Dept': pd.DataFrame({
            'Department': ['Computer Science', 'Computer Science', 'Computer Science',
                          'Engineering', 'Engineering', 'Engineering',
                          'Business', 'Business', 'Business'],
            'Resource': ['Python Programming', 'Data Structures', 'AI Fundamentals',
                        'Thermodynamics', 'Circuit Analysis', 'Fluid Mechanics',
                        'Financial Accounting', 'Marketing Strategy', 'Business Ethics'],
            'Usage_Count': [345, 278, 234, 198, 167, 145, 189, 156, 123],
            'Category': ['Textbook', 'Textbook', 'Reference', 'Textbook', 'Reference', 'Textbook',
                        'Textbook', 'Reference', 'Textbook']
        })
    }
    
    # 3. OPERATIONAL DASHBOARD DATA
    operational_data = {
        'Daily_Metrics': pd.DataFrame({
            'Date': [(today - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(14, 0, -1)],
            'Day': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']*2,
            'Checkouts': [45, 52, 48, 56, 51, 29, 18, 47, 53, 58, 52, 49, 31, 22],
            'Returns': [42, 48, 45, 52, 47, 25, 15, 44, 49, 54, 48, 46, 28, 19],
            'New_Registrations': [5, 7, 6, 8, 7, 3, 2, 6, 7, 8, 7, 6, 4, 3],
            'Room_Occupancy_%': [78, 85, 82, 88, 84, 45, 32, 82, 86, 90, 85, 83, 52, 38]
        }),
        
        'Room_Status': pd.DataFrame({
            'Room': ['R101', 'R102', 'R103', 'R104', 'R105', 'R201', 'R202', 'R203'],
            'Type': ['Study', 'Group', 'Study', 'Group', 'Study', 'Group', 'Study', 'Group'],
            'Capacity': [4, 8, 4, 8, 4, 10, 4, 8],
            'Today_Bookings': [4, 3, 5, 2, 4, 3, 4, 2],
            'Hours_Booked': [8.5, 6.0, 10.0, 4.5, 8.0, 6.0, 8.0, 4.0],
            'Status': ['Fully Booked', 'Available', 'Fully Booked', 'Available', 
                      'Fully Booked', 'Available', 'Fully Booked', 'Available'],
            'Next_Available': ['Tomorrow', 'Now', 'Tomorrow', 'Now', 'Tomorrow', 'Now', 'Tomorrow', 'Now']
        }),
        
        'Overdue_Items': pd.DataFrame({
            'Item_ID': ['BK-001', 'BK-002', 'BK-003', 'BK-004', 'BK-005', 'BK-006', 'BK-007', 'BK-008'],
            'Title': ['Introduction to Algorithms', 'Thermodynamics', 'Financial Management',
                     'Artificial Intelligence', 'Circuit Analysis', 'Marketing Principles',
                     'Data Structures', 'Business Ethics'],
            'Borrower_ID': ['STU-2024-001', 'STU-2024-002', 'STU-2024-003', 'STU-2024-004',
                           'STU-2024-005', 'STU-2024-006', 'STU-2024-007', 'STU-2024-008'],
            'Days_Overdue': [5, 12, 3, 8, 15, 2, 7, 4],
            'Category': ['Textbook', 'Reference', 'Textbook', 'Reference', 
                        'Textbook', 'Reference', 'Textbook', 'Reference'],
            'Fine_Amount': [2.50, 6.00, 1.50, 4.00, 7.50, 1.00, 3.50, 2.00]
        }),
        
        'System_Metrics': pd.DataFrame({
            'Metric': ['Server Uptime (%)', 'Database Response (ms)', 'Wi-Fi Users', 
                      'Printer Queue', 'Website Visits', 'App Logins', 'API Calls'],
            'Value': [99.8, 45, 156, 3, 2345, 567, 12345],
            'Threshold': [99.5, 100, 200, 10, 2000, 500, 10000],
            'Status': ['Normal', 'Normal', 'Normal', 'Normal', 'High', 'High', 'High']
        })
    }
    
    return executive_data, department_data, operational_data

# ==================== EXCEL STYLING FUNCTIONS ====================
def apply_cell_styling(cell, style_type, palette):
    """Apply consistent styling to cells"""
    
    if style_type == 'title':
        cell.font = Font(name='Calibri', size=18, bold=True, color=palette['text'])
        cell.fill = PatternFill(start_color=palette['primary'], end_color=palette['primary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    elif style_type == 'header':
        cell.font = Font(name='Calibri', size=11, bold=True, color=palette['text'])
        cell.fill = PatternFill(start_color=palette['secondary'], end_color=palette['secondary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
    elif style_type == 'data':
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
    elif style_type == 'kpi_good':
        cell.font = Font(name='Calibri', size=10, bold=True, color='107C10')
        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
    elif style_type == 'kpi_warning':
        cell.font = Font(name='Calibri', size=10, bold=True, color='9C6500')
        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
    elif style_type == 'kpi_critical':
        cell.font = Font(name='Calibri', size=10, bold=True, color='9C0006')
        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

def create_dashboard_workbook(dashboard_type, title, data_dict):
    """Create a complete Excel dashboard workbook"""
    
    palette = COLOR_PALETTES[dashboard_type]
    wb = Workbook()
    
    # Remove default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)
    
    for sheet_name, df in data_dict.items():
        ws = wb.create_sheet(sheet_name.replace('_', ' '))
        
        # Add title row
        ws.merge_cells('A1:Z1')
        title_cell = ws['A1']
        title_cell.value = f"{title} - {sheet_name.replace('_', ' ')}"
        apply_cell_styling(title_cell, 'title', palette)
        
        # Add timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(name='Calibri', size=9, italic=True, color='666666')
        
        # Write data starting from row 4
        start_row = 4
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == start_row:  # Header row
                    apply_cell_styling(cell, 'header', palette)
                else:  # Data row
                    apply_cell_styling(cell, 'data', palette)
                    
                    # Apply conditional formatting for status indicators
                    if isinstance(value, str):
                        if 'Above' in value or 'Good' in value or 'Normal' in value:
                            apply_cell_styling(cell, 'kpi_good', palette)
                        elif 'Warning' in value or 'Medium' in value:
                            apply_cell_styling(cell, 'kpi_warning', palette)
                        elif 'Critical' in value or 'Below' in value:
                            apply_cell_styling(cell, 'kpi_critical', palette)
        
        # Auto-adjust column widths
        for col_idx in range(1, df.shape[1] + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(start_row, len(df) + start_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes for scrolling
        ws.freeze_panes = ws[f'A{start_row + 1}']
        
        # Add filter to header row
        ws.auto_filter.ref = ws.dimensions
    
    return wb

def add_charts_to_workbook(wb, dashboard_type):
    """Add charts to the workbook based on dashboard type"""
    
    if dashboard_type == 'executive':
        # Monthly Trends Line Chart
        if 'Monthly Trends' in wb.sheetnames:
            ws = wb['Monthly Trends']
            chart = LineChart()
            chart.title = "Monthly Library Usage Trends"
            chart.style = 13
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Month"
            chart.legend.position = 'b'
            
            data = Reference(ws, min_col=2, min_row=4, max_col=5, max_row=15)
            categories = Reference(ws, min_col=1, min_row=5, max_row=15)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.width = 20
            chart.height = 10
            ws.add_chart(chart, "G3")
        
        # Resource Categories Pie Chart
        if 'Resource Categories' in wb.sheetnames:
            ws = wb['Resource Categories']
            chart = PieChart()
            chart.title = "Resource Usage Distribution"
            
            data = Reference(ws, min_col=2, min_row=4, max_row=11)
            labels = Reference(ws, min_col=1, min_row=5, max_row=11)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            
            # Add data labels
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            
            chart.width = 12
            chart.height = 8
            ws.add_chart(chart, "E3")
    
    elif dashboard_type == 'department':
        # Department Comparison Bar Chart
        if 'Department Comparison' in wb.sheetnames:
            ws = wb['Department Comparison']
            chart = BarChart()
            chart.title = "Department Performance Comparison"
            chart.type = "col"
            chart.style = 10
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Department"
            chart.legend.position = 'b'
            
            data = Reference(ws, min_col=2, min_row=4, max_col=5, max_row=11)
            categories = Reference(ws, min_col=1, min_row=5, max_row=11)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.width = 22
            chart.height = 12
            ws.add_chart(chart, "H3")
    
    elif dashboard_type == 'operational':
        # Daily Metrics Line Chart
        if 'Daily Metrics' in wb.sheetnames:
            ws = wb['Daily Metrics']
            chart = LineChart()
            chart.title = "Daily Operational Trends (14 Days)"
            chart.style = 12
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Date"
            chart.legend.position = 'b'
            
            data = Reference(ws, min_col=3, min_row=4, max_col=6, max_row=17)
            categories = Reference(ws, min_col=1, min_row=5, max_row=17)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.width = 22
            chart.height = 12
            ws.add_chart(chart, "I3")
    
    return wb

# ==================== MAIN FUNCTION ====================
def main():
    """Main function to generate all dashboards"""
    
    print("=" * 70)
    print("EXCEL DASHBOARDS GENERATOR - LIBRARY ANALYTICS")
    print("=" * 70)
    
    # Create output folder
    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)
        print(f" Created output folder: {os.path.abspath(OUTPUT_FOLDER)}")
    
    print("\n Generating library data...")
    executive_data, department_data, operational_data = generate_library_data()
    
    # 1. Create Executive Dashboard
    print("\n" + "-" * 50)
    print("1. Creating EXECUTIVE DASHBOARD...")
    exec_wb = create_dashboard_workbook(
        'executive',
        DASHBOARD_TITLES['executive'],
        executive_data
    )
    exec_wb = add_charts_to_workbook(exec_wb, 'executive')
    
    exec_path = os.path.join(OUTPUT_FOLDER, 'Executive_Dashboard.xlsx')
    exec_wb.save(exec_path)
    print(f"   Saved: {exec_path}")
    print(f"    Sheets: {', '.join(exec_wb.sheetnames)}")
    
    # 2. Create Department Dashboard
    print("\n2. Creating DEPARTMENT DASHBOARD...")
    dept_wb = create_dashboard_workbook(
        'department',
        DASHBOARD_TITLES['department'],
        department_data
    )
    dept_wb = add_charts_to_workbook(dept_wb, 'department')
    
    dept_path = os.path.join(OUTPUT_FOLDER, 'Department_Dashboard.xlsx')
    dept_wb.save(dept_path)
    print(f"    Saved: {dept_path}")
    print(f"    Sheets: {', '.join(dept_wb.sheetnames)}")
    
    # 3. Create Operational Dashboard
    print("\n3. Creating OPERATIONAL DASHBOARD...")
    ops_wb = create_dashboard_workbook(
        'operational',
        DASHBOARD_TITLES['operational'],
        operational_data
    )
    ops_wb = add_charts_to_workbook(ops_wb, 'operational')
    
    ops_path = os.path.join(OUTPUT_FOLDER, 'Operational_Dashboard.xlsx')
    ops_wb.save(ops_path)
    print(f"   Saved: {ops_path}")
    print(f"    Sheets: {', '.join(ops_wb.sheetnames)}")
    
    # Summary
    print("\n" + "=" * 70)
    print(" DASHBOARD GENERATION COMPLETE!")
    print("=" * 70)
    
    print(f"\n Output Location: {os.path.abspath(OUTPUT_FOLDER)}")
    print("\n Generated Files:")
    print(f"   1. Executive_Dashboard.xlsx")
    print(f"      - KPI Summary")
    print(f"      - Monthly Trends")
    print(f"      - Resource Categories")
    print(f"      - Peak Hours Analysis")
    
    print(f"\n   2. Department_Dashboard.xlsx")
    print(f"      - Department Comparison")
    print(f"      - Department Trends")
    print(f"      - Top Resources by Department")
    
    print(f"\n   3. Operational_Dashboard.xlsx")
    print(f"      - Daily Metrics")
    print(f"      - Room Status")
    print(f"      - Overdue Items")
    print(f"      - System Metrics")
    
    print("\nTotal Sheets Created: 11")
    print(" Total Charts Created: 4")
    print(f" Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    print("\n" + "=" * 70)
    print(" TIPS:")
    print("   • Open files in Excel for interactive filtering")
    print("   • Use the auto-filters in header rows")
    print("   • Charts update automatically when data changes")
    print("   • Connect to your database for live data (see comments)")
    print("=" * 70)

# ==================== DATABASE CONNECTION EXAMPLE ====================


# ==================== ENTRY POINT ====================
if __name__ == "__main__":
    try:
        main()
        
        # Optional: Open the folder in Windows Explorer
        open_folder = input("\nDo you want to open the output folder? (y/n): ")
        if open_folder.lower() == 'y':
            os.startfile(os.path.abspath(OUTPUT_FOLDER))
            
    except Exception as e:
        print(f"\n❌ Error: {e}")
        print("\n🔧 Troubleshooting:")
        print("   1. Make sure you have required packages: pip install pandas openpyxl")
        print("   2. Check if you have write permissions in the folder")
        print("   3. Close any Excel files that might be open")
        
        # Install missing packages
        install_packages = input("\nInstall missing packages? (y/n): ")
        if install_packages.lower() == 'y':
            os.system("pip install pandas openpyxl")
            print("\nPackages installed. Please run the script again.")